import json
import os

from openpyxl import  load_workbook

def get_parent_framework_path():
    current_path = os.getcwd()
    get_parent = os.path.dirname(current_path)
    return get_parent

def get_data_from_inputs(data_key):
    file_path = os.path.join(get_parent_framework_path(), "data", "input_data.json")
    with open(file_path, "r") as file:
        data = json.load(file)
        return data[data_key]

def read_excel_data(sheet_name):
    excel_file = os.path.join(get_parent_framework_path(), "data", "testdata.xlsx")
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col= ws.max_column):
        data.append(tuple([cell.value for cell in row]))
    wb.close()
    return data

def read_csv_data(sheet_name):
    excel_file = os.path.join(get_parent_framework_path(), "data", "testdata.xlsx")
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col= ws.max_column):
        data.append(tuple([cell.value for cell in row]))
    wb.close()
    return data
